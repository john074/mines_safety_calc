import openpyxl

def parse_excel_sheet(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не найден в файле.")
    ws = wb[sheet_name]

    result = {}
    current_factor = None

    for row in ws.iter_rows(min_row=7, min_col=3, max_col=5):
        factor_cell, coefficient_cell, criterion_cell = row

        if factor_cell.value:
            current_factor = factor_cell.value.strip()
            result[current_factor] = {}

        if current_factor is not None and coefficient_cell.value is not None and criterion_cell.value is not None:
            try:
                coefficient = float(coefficient_cell.value)
            except ValueError:
                continue 
            criterion = str(criterion_cell.value).strip()
            result[current_factor][criterion] = coefficient

    return result

def print_dict_as_code(data_dict):
    import pprint
    pprint.pprint(data_dict, width=120, sort_dicts=False)

if __name__ == "__main__":
    excel_file = "Оценка уровня опасности.xlsx"
    sheet = input("Введите название листа: ")

    try:
        data = parse_excel_sheet(excel_file, sheet)
        print_dict_as_code(data)
    except Exception as e:
        print(f"Ошибка: {e}")
